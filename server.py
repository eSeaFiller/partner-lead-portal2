#!/usr/bin/env python3
import warnings

warnings.simplefilter("ignore", DeprecationWarning)

import cgi
import json
import os
import re
import shutil
import uuid
from difflib import SequenceMatcher
from io import BytesIO
from datetime import datetime, timezone
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, urlparse

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
STATIC_DIR = BASE_DIR / "static"
DEFAULT_RUNTIME_DIR = Path("/tmp/partner-lead-portal") if os.environ.get("VERCEL") else BASE_DIR
DATA_DIR = Path(os.environ.get("LEAD_DATA_DIR", DEFAULT_RUNTIME_DIR / "data"))
EXPORT_DIR = Path(os.environ.get("LEAD_EXPORT_DIR", DEFAULT_RUNTIME_DIR / "exports"))
DATA_FILE = DATA_DIR / "leads.json"
TEMPLATE_PATH = Path(
    os.environ.get(
        "LEAD_TEMPLATE_PATH",
        BASE_DIR / "templates" / "oversea_lead_import_template.xlsx",
    )
)
ADMIN_KEY = os.environ.get("ADMIN_KEY", "lark-admin")
ADMIN_AUTH_DISABLED = os.environ.get("ADMIN_AUTH_DISABLED", "1") == "1"
PUBLIC_ONLY = os.environ.get("PUBLIC_ONLY") == "1"
CORS_ALLOW_ORIGIN = os.environ.get("CORS_ALLOW_ORIGIN", "")

FIELD_TO_HEADER = {
    "firstName": "First Name",
    "lastName": "Last Name",
    "country": "Country",
    "companyName": "Company Name",
    "mobileNumber": "Mobile Number",
    "workEmail": "Work Email",
    "jobTitle": "Job Title",
    "companySize": "Company Size",
    "industry": "Industry",
    "subIndustry": "Sub Industry",
    "trackingCode": "Tracking Code",
}

FIELD_LABELS = {
    **FIELD_TO_HEADER,
    "companyName": "Company Name",
    "workEmail": "Work Email",
}

FIELD_ALIASES = {
    "firstName": ["first", "given name", "given", "first_name", "firstname", "名"],
    "lastName": ["last", "family name", "surname", "last_name", "lastname", "姓"],
    "country": ["country code", "country/region", "region", "market", "国家", "国家/地区"],
    "companyName": ["company", "account", "organization", "organisation", "company_name", "公司", "公司名称"],
    "mobileNumber": ["mobile", "phone", "phone number", "mobile phone", "tel", "telephone", "手机号", "电话"],
    "workEmail": ["email", "work mail", "business email", "work_email", "mail", "邮箱", "工作邮箱"],
    "jobTitle": ["title", "role", "position", "job", "job_title", "职位", "职务"],
    "companySize": ["size", "employee size", "employees", "headcount", "company_size", "公司规模"],
    "industry": ["sector", "vertical", "行业"],
    "subIndustry": ["subindustry", "sub industry", "sub-sector", "sub sector", "子行业"],
    "trackingCode": ["tracking", "campaign", "campaign code", "tracking_code", "活动代码"],
}

JOB_TITLE_ALIASES = {
    "chief executive officer": "CEO",
    "co founder": "CEO",
    "cofounder": "CEO",
    "founder": "CEO",
    "owner": "CEO",
    "general manager": "GM",
    "managing director": "GM",
    "vice president": "VP",
    "svp": "VP",
    "evp": "VP",
    "chief strategy officer": "ChiefStrategyOfficer",
    "strategy officer": "ChiefStrategyOfficer",
    "board secretary": "BoardSecretary",
    "chief technology officer": "CTO",
    "chief information officer": "CIO",
    "chief financial officer": "CFO",
    "chief marketing officer": "CMO",
    "chief operating officer": "COO",
    "chief human resources officer": "CHO",
    "chief people officer": "CHO",
    "chief security officer": "CISO",
    "chief information security officer": "CISO",
    "chief digital officer": "CDO",
    "chief data officer": "CDO",
    "chief experience officer": "CXO",
    "it manager": "ITManager",
    "information technology manager": "ITManager",
    "hr manager": "HRSupervisor",
    "human resources manager": "HRSupervisor",
    "sales manager": "SalesSupervisor",
    "sales supervisor": "SalesSupervisor",
    "admin manager": "AdministrativeSupervisor",
    "administrative manager": "AdministrativeSupervisor",
    "product manager": "Product",
    "operations manager": "OperationsManager",
    "operation manager": "OperationsManager",
    "technical manager": "TechnicalManager",
    "technology manager": "TechnicalManager",
    "marketing manager": "MarketingManager",
    "finance manager": "FinancialOfficer",
    "financial manager": "FinancialOfficer",
    "accounting manager": "FinancialOfficer",
    "procurement manager": "PurchasingSupervisor",
    "purchasing manager": "PurchasingSupervisor",
    "business development manager": "BusinessSupervisor",
    "business manager": "BusinessSupervisor",
    "design manager": "Design",
    "designer": "Design",
    "head of department": "HeadofDep",
    "department head": "HeadofDep",
    "manager": "Manager",
    "director": "Director",
    "engineer": "Engineering",
    "software engineer": "Engineering",
    "developer": "Engineering",
    "professor": "Professor",
    "teacher": "Professor",
    "employee": "Employee",
    "staff": "Employee",
    "student": "Student",
    "others": "Others",
    "other": "Others",
}

JOB_TITLE_KEYWORD_RULES = [
    (r"\bchief\b.*\bexecutive\b|\bceo\b|\bfounder\b|\bowner\b", "CEO"),
    (r"\bgeneral manager\b|\bmanaging director\b|\bcountry manager\b", "GM"),
    (r"\bvice president\b|\bsvp\b|\bevp\b|\bvp\b", "VP"),
    (r"\bchief\b.*\bstrategy\b|\bstrategy officer\b", "ChiefStrategyOfficer"),
    (r"\bboard secretary\b", "BoardSecretary"),
    (r"\bchief\b.*\boperating\b|\bcoo\b", "COO"),
    (r"\bchief\b.*\binformation\b|\bcio\b", "CIO"),
    (r"\bchief\b.*\btechnology\b|\bcto\b", "CTO"),
    (r"\bchief\b.*\bsecurity\b|\bciso\b", "CISO"),
    (r"\bchief\b.*\bmarketing\b|\bcmo\b", "CMO"),
    (r"\bchief\b.*\bfinancial\b|\bcfo\b", "CFO"),
    (r"\bchief\b.*\bdata\b|\bchief\b.*\bdigital\b|\bcdo\b", "CDO"),
    (r"\bchief\b.*\bpeople\b|\bchief\b.*\bhuman\b|\bcho\b", "CHO"),
    (r"\bchief\b|\bcxo\b", "CXO"),
    (r"\bit\b.*\bmanager\b|\binformation technology\b", "ITManager"),
    (r"\bhr\b|\bhuman resources\b|\bpeople ops\b", "HRSupervisor"),
    (r"\bsales\b", "SalesSupervisor"),
    (r"\badmin\b|\badministrative\b", "AdministrativeSupervisor"),
    (r"\bproduct\b", "Product"),
    (r"\boperations?\b", "OperationsManager"),
    (r"\btechnical\b|\btechnology\b", "TechnicalManager"),
    (r"\bmarketing\b", "MarketingManager"),
    (r"\bfinance\b|\bfinancial\b|\baccounting\b", "FinancialOfficer"),
    (r"\bprocurement\b|\bpurchasing\b", "PurchasingSupervisor"),
    (r"\bbusiness development\b|\bbd\b", "BusinessSupervisor"),
    (r"\bdesign\b|\bdesigner\b", "Design"),
    (r"\bhead\b.*\bdepartment\b|\bdepartment head\b", "HeadofDep"),
    (r"\bdirector\b", "Director"),
    (r"\bengineer\b|\bdeveloper\b", "Engineering"),
    (r"\bprofessor\b|\bteacher\b", "Professor"),
    (r"\bstudent\b|\bintern\b", "Student"),
    (r"\bmanager\b|\bsupervisor\b|\blead\b", "Manager"),
    (r"\bstaff\b|\bemployee\b|\bexecutive\b|\bassistant\b|\bofficer\b", "Employee"),
]

INDUSTRY_ALIASES = {
    "saas": ("Technology & Internet", "Software & SaaS"),
    "tech": ("Technology & Internet", "Technology & Internet Others"),
    "technology": ("Technology & Internet", "Technology & Internet Others"),
    "internet": ("Technology & Internet", "Technology & Internet Others"),
    "software": ("Technology & Internet", "Software & SaaS"),
    "cloud software": ("Technology & Internet", "Software & SaaS"),
    "enterprise software": ("Technology & Internet", "Software & SaaS"),
    "it": ("Technology & Internet", "IT Infrastructure"),
    "it services": ("Technology & Internet", "IT Infrastructure"),
    "information technology": ("Technology & Internet", "IT Infrastructure"),
    "ai": ("Technology & Internet", "AI / Big Data / IoT / Robotics / VR"),
    "artificial intelligence": ("Technology & Internet", "AI / Big Data / IoT / Robotics / VR"),
    "big data": ("Technology & Internet", "AI / Big Data / IoT / Robotics / VR"),
    "iot": ("Technology & Internet", "AI / Big Data / IoT / Robotics / VR"),
    "robotics": ("Technology & Internet", "AI / Big Data / IoT / Robotics / VR"),
    "ecommerce": ("Technology & Internet", "Ecommerce"),
    "e-commerce": ("Technology & Internet", "Ecommerce"),
    "online marketplace": ("Technology & Internet", "Specialized Online Marketplace"),
    "gaming": ("Technology & Internet", "Gaming"),
    "game": ("Technology & Internet", "Gaming"),
    "social media": ("Technology & Internet", "Social Media, News & Entertainment"),
    "advertising": ("Technology & Internet", "Marketing, Growth & Advertising"),
    "martech": ("Technology & Internet", "Marketing, Growth & Advertising"),
    "marketing agency": ("Technology & Internet", "Marketing, Growth & Advertising"),
    "telecom": ("Telecommunications", "Telecommunications Others"),
    "telco": ("Telecommunications", "Telecommunications Others"),
    "education": ("Education", "Education Others"),
    "edtech": ("Education", "Online Education"),
    "online education": ("Education", "Online Education"),
    "school": ("Education", "K-12 Education"),
    "university": ("Education", "University & College"),
    "finance": ("Financial Services", "Financial Services Others"),
    "banking": ("Financial Services", "Banks & Asset Management"),
    "bank": ("Financial Services", "Banks & Asset Management"),
    "fintech": ("Financial Services", "FinTech"),
    "insurance": ("Financial Services", "Insurance"),
    "vc": ("Financial Services", "PE & VC"),
    "venture capital": ("Financial Services", "PE & VC"),
    "healthcare": ("Healthcare", "Healthcare Others"),
    "health care": ("Healthcare", "Healthcare Others"),
    "healthtech": ("Healthcare", "Health Tech"),
    "hospital": ("Healthcare", "Hospitals, Clinics & Specialists"),
    "pharma": ("Healthcare", "Pharmaceuticals & Biotech"),
    "biotech": ("Healthcare", "Pharmaceuticals & Biotech"),
    "medical device": ("Healthcare", "Medical Devices & Equipment"),
    "retail": ("Consumer, Retail & Wholesale", "Consumer Retail Products & Brands"),
    "consumer goods": ("Consumer, Retail & Wholesale", "Consumer Retail Products & Brands"),
    "fmcg": ("Consumer, Retail & Wholesale", "Consumer Retail Products & Brands"),
    "hotel": ("Consumer, Retail & Wholesale", "Accommodation & Hotels"),
    "hospitality": ("Consumer, Retail & Wholesale", "Accommodation & Hotels"),
    "food and beverage": ("Consumer, Retail & Wholesale", "Food & Beverage"),
    "f&b": ("Consumer, Retail & Wholesale", "Food & Beverage"),
    "wholesale": ("Consumer, Retail & Wholesale", "Wholesale"),
    "manufacturing": ("Industrial Manufacturing", "Industrial Manufacturing Others"),
    "manufacturer": ("Industrial Manufacturing", "Industrial Manufacturing Others"),
    "electronics": ("Industrial Manufacturing", "Electronics & Telecommunications Equipment"),
    "automotive": ("Industrial Manufacturing", "Vehicles & Automobile Parts"),
    "automobile": ("Industrial Manufacturing", "Vehicles & Automobile Parts"),
    "chemical": ("Industrial Manufacturing", "Chemicals"),
    "professional services": ("Professional Services", "Professional Services Others"),
    "consulting": ("Professional Services", "Management Consulting"),
    "management consulting": ("Professional Services", "Management Consulting"),
    "legal": ("Professional Services", "Legal"),
    "law firm": ("Professional Services", "Legal"),
    "accounting": ("Professional Services", "Accounting, Audit & Tax"),
    "audit": ("Professional Services", "Accounting, Audit & Tax"),
    "staffing": ("Professional Services", "HR & Staffing"),
    "real estate": ("Real Estate & Construction", "Commercial & Residential Real Estate"),
    "property": ("Real Estate & Construction", "Commercial & Residential Real Estate"),
    "construction": ("Real Estate & Construction", "Construction"),
    "logistics": ("Transportation & Logistics", "Freight & Logistics Services"),
    "transportation": ("Transportation & Logistics", "Transportation & Logistics Others"),
    "shipping": ("Transportation & Logistics", "Marine Shipping & Transportation"),
    "airline": ("Transportation & Logistics", "Airlines"),
    "government": ("Public Service & Non-profits", "Government"),
    "ngo": ("Public Service & Non-profits", "NGOs"),
    "nonprofit": ("Public Service & Non-profits", "NGOs"),
    "energy": ("Energy, Mining, Utilities & Waste", "Energy, Mining, Utilities & Waste Others"),
    "oil and gas": ("Energy, Mining, Utilities & Waste", "Oil & Gas"),
    "mining": ("Energy, Mining, Utilities & Waste", "Minerals & Mining"),
    "agriculture": ("Agriculture", "Agriculture Others"),
    "crypto": ("Web3, Blockchain & Crypto", "Web3, Blockchain & Crypto"),
    "blockchain": ("Web3, Blockchain & Crypto", "Web3, Blockchain & Crypto"),
    "web3": ("Web3, Blockchain & Crypto", "Web3, Blockchain & Crypto"),
}

INDUSTRY_KEYWORD_RULES = [
    (r"\bsaas\b|\bsoftware\b|\bcrm\b|\berp\b", ("Technology & Internet", "Software & SaaS")),
    (r"\bai\b|\bartificial intelligence\b|\bbig data\b|\biot\b|\brobotics\b|\bvr\b", ("Technology & Internet", "AI / Big Data / IoT / Robotics / VR")),
    (r"\be-?commerce\b|\bonline shop\b|\bonline store\b", ("Technology & Internet", "Ecommerce")),
    (r"\bmarketplace\b", ("Technology & Internet", "Specialized Online Marketplace")),
    (r"\bgaming\b|\bgame\b", ("Technology & Internet", "Gaming")),
    (r"\bfintech\b", ("Financial Services", "FinTech")),
    (r"\bbank\b|\bbanking\b", ("Financial Services", "Banks & Asset Management")),
    (r"\binsurance\b", ("Financial Services", "Insurance")),
    (r"\bhealth ?tech\b", ("Healthcare", "Health Tech")),
    (r"\bhospital\b|\bclinic\b", ("Healthcare", "Hospitals, Clinics & Specialists")),
    (r"\bpharma\b|\bbiotech\b", ("Healthcare", "Pharmaceuticals & Biotech")),
    (r"\bedtech\b|\bonline education\b", ("Education", "Online Education")),
    (r"\bschool\b|\bk-?12\b", ("Education", "K-12 Education")),
    (r"\buniversity\b|\bcollege\b", ("Education", "University & College")),
    (r"\bretail\b|\bfmcg\b|\bconsumer goods\b", ("Consumer, Retail & Wholesale", "Consumer Retail Products & Brands")),
    (r"\bhotel\b|\bhospitality\b", ("Consumer, Retail & Wholesale", "Accommodation & Hotels")),
    (r"\bf&b\b|\bfood\b|\bbeverage\b", ("Consumer, Retail & Wholesale", "Food & Beverage")),
    (r"\bmanufactur", ("Industrial Manufacturing", "Industrial Manufacturing Others")),
    (r"\bconsulting\b", ("Professional Services", "Management Consulting")),
    (r"\blegal\b|\blaw firm\b", ("Professional Services", "Legal")),
    (r"\breal estate\b|\bproperty\b", ("Real Estate & Construction", "Commercial & Residential Real Estate")),
    (r"\blogistics\b|\bfreight\b", ("Transportation & Logistics", "Freight & Logistics Services")),
    (r"\bcrypto\b|\bblockchain\b|\bweb3\b", ("Web3, Blockchain & Crypto", "Web3, Blockchain & Crypto")),
]

REQUIRED_FIELDS = [field for field in FIELD_TO_HEADER if field != "trackingCode"]

EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
MAX_UPLOAD_BYTES = 8 * 1024 * 1024


def normalize_lookup_key(value):
    return re.sub(r"[\W_]+", "", str(value or "").strip().lower(), flags=re.UNICODE)


def normalize_header(value):
    return normalize_lookup_key(value)


def utc_now():
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def normalize_company(value):
    value = (value or "").lower()
    value = re.sub(r"[^a-z0-9]+", " ", value)
    suffixes = {
        "inc",
        "ltd",
        "limited",
        "co",
        "company",
        "corp",
        "corporation",
        "pte",
        "plc",
        "llc",
        "gmbh",
        "sdn",
        "bhd",
    }
    return " ".join(part for part in value.split() if part not in suffixes)


def read_column_values(ws, column, start_row=2):
    values = []
    for row in range(start_row, ws.max_row + 1):
        value = ws[f"{column}{row}"].value
        if value not in (None, ""):
            values.append(str(value).strip())
    return values


def load_schema():
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"Template not found: {TEMPLATE_PATH}")

    # This template stores large enum ranges in a way that openpyxl can
    # under-report in read-only mode, so load normally and keep the schema exact.
    wb = load_workbook(TEMPLATE_PATH, read_only=False, data_only=True)
    country_ws = wb["Country Enums"]
    job_ws = wb["Job Title Enums"]
    size_ws = wb["Company Size Enums"]
    industry_ws = wb["Industry & Sub Industry Enums"]

    countries = [
        {
            "value": str(country_ws[f"A{row}"].value).strip(),
            "label": str(country_ws[f"C{row}"].value or country_ws[f"B{row}"].value or "").strip(),
        }
        for row in range(2, country_ws.max_row + 1)
        if country_ws[f"A{row}"].value
    ]

    job_titles = [
        {
            "value": str(job_ws[f"A{row}"].value).strip(),
            "label": str(job_ws[f"B{row}"].value or job_ws[f"A{row}"].value).strip(),
        }
        for row in range(2, job_ws.max_row + 1)
        if job_ws[f"A{row}"].value
    ]

    company_sizes = read_column_values(size_ws, "A")

    industries = {}
    for row in range(2, industry_ws.max_row + 1):
        industry = industry_ws[f"A{row}"].value
        sub_industry = industry_ws[f"B{row}"].value
        industry_label = industry_ws[f"C{row}"].value
        sub_label = industry_ws[f"D{row}"].value
        if not industry or not sub_industry:
            continue
        industry = str(industry).strip()
        if industry not in industries:
            industries[industry] = {
                "value": industry,
                "label": str(industry_label or industry).strip(),
                "subIndustries": [],
            }
        industries[industry]["subIndustries"].append(
            {
                "value": str(sub_industry).strip(),
                "label": str(sub_label or sub_industry).strip(),
            }
        )

    return {
        "templatePath": str(TEMPLATE_PATH),
        "fields": FIELD_TO_HEADER,
        "requiredFields": REQUIRED_FIELDS,
        "countries": countries,
        "jobTitles": job_titles,
        "companySizes": company_sizes,
        "industries": list(industries.values()),
    }


SCHEMA = load_schema()


def load_leads():
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    if not DATA_FILE.exists():
        return []
    with DATA_FILE.open("r", encoding="utf-8") as f:
        return json.load(f)


def save_leads(leads):
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    temp = DATA_FILE.with_suffix(".tmp")
    with temp.open("w", encoding="utf-8") as f:
        json.dump(leads, f, indent=2, ensure_ascii=False)
    temp.replace(DATA_FILE)


def schema_sets():
    countries = {item["value"] for item in SCHEMA["countries"]}
    job_titles = {item["value"] for item in SCHEMA["jobTitles"]}
    company_sizes = set(SCHEMA["companySizes"])
    industries = {item["value"]: {sub["value"] for sub in item["subIndustries"]} for item in SCHEMA["industries"]}
    return countries, job_titles, company_sizes, industries


def enum_lookup_maps():
    country_map = {}
    for item in SCHEMA["countries"]:
        country_map[normalize_lookup_key(item["value"])] = item["value"]
        country_map[normalize_lookup_key(item["label"])] = item["value"]

    job_map = {}
    for item in SCHEMA["jobTitles"]:
        job_map[normalize_lookup_key(item["value"])] = item["value"]
        job_map[normalize_lookup_key(item["label"])] = item["value"]
    for alias, value in JOB_TITLE_ALIASES.items():
        job_map[normalize_lookup_key(alias)] = value

    size_map = {normalize_lookup_key(size): size for size in SCHEMA["companySizes"]}
    size_map.update(
        {
            normalize_lookup_key("1 to 10"): "1-10",
            normalize_lookup_key("11 to 20"): "11-20",
            normalize_lookup_key("21 to 50"): "21-50",
            normalize_lookup_key("51 to 100"): "51-100",
            normalize_lookup_key("101 to 250"): "101-250",
            normalize_lookup_key("251 to 500"): "251-500",
            normalize_lookup_key("501 to 999"): "501-999",
            normalize_lookup_key("1000 to 4999"): "1000-4999",
            normalize_lookup_key("5000 to 9999"): "5000-9999",
            normalize_lookup_key("10000+"): ">10000",
            normalize_lookup_key("> 10000"): ">10000",
            normalize_lookup_key("more than 10000"): ">10000",
        }
    )

    industry_map = {}
    subindustry_map = {}
    subindustry_by_industry = {}
    subindustry_parent = {}
    for industry in SCHEMA["industries"]:
        industry_map[normalize_lookup_key(industry["value"])] = industry["value"]
        industry_map[normalize_lookup_key(industry["label"])] = industry["value"]
        subindustry_by_industry[industry["value"]] = {}
        for sub in industry["subIndustries"]:
            subindustry_map[normalize_lookup_key(sub["value"])] = sub["value"]
            subindustry_map[normalize_lookup_key(sub["label"])] = sub["value"]
            subindustry_parent[sub["value"]] = industry["value"]
            subindustry_by_industry[industry["value"]][normalize_lookup_key(sub["value"])] = sub["value"]
            subindustry_by_industry[industry["value"]][normalize_lookup_key(sub["label"])] = sub["value"]

    return {
        "country": country_map,
        "jobTitle": job_map,
        "companySize": size_map,
        "industry": industry_map,
        "subIndustry": subindustry_map,
        "subIndustryByIndustry": subindustry_by_industry,
        "subIndustryParent": subindustry_parent,
    }


def map_enum_value(field, value, maps, warnings):
    original = str(value or "").strip()
    if not original:
        return ""
    mapped = maps[field].get(normalize_lookup_key(original))
    if mapped:
        if mapped != original:
            warnings[field] = f"Cleaned '{original}' to '{mapped}'"
        return mapped
    return original


def company_size_bucket(number):
    buckets = [
        (1, 10, "1-10"),
        (11, 20, "11-20"),
        (21, 50, "21-50"),
        (51, 100, "51-100"),
        (101, 250, "101-250"),
        (251, 500, "251-500"),
        (501, 999, "501-999"),
        (1000, 4999, "1000-4999"),
        (5000, 9999, "5000-9999"),
    ]
    for low, high, label in buckets:
        if low <= number <= high:
            return label
    if number >= 10000:
        return ">10000"
    return ""


def clean_company_size(value, maps):
    original = str(value or "").strip()
    if not original:
        return "", ""

    mapped = maps["companySize"].get(normalize_lookup_key(original))
    if mapped:
        return mapped, "exact"

    lowered = original.lower().replace(",", "")
    if re.search(r"\b(above|over|more than|greater than)\b", lowered) and re.search(r"\b10000\b", lowered):
        return ">10000", "range"

    numbers = [int(match) for match in re.findall(r"\d+", lowered)]
    if not numbers:
        return original, ""

    if re.search(r"\b(less than|under|below|fewer than|up to)\b", lowered):
        return company_size_bucket(max(1, numbers[0])), "range"

    if re.search(r"\b(above|over|more than|greater than)\b", lowered):
        return company_size_bucket(numbers[0] + 1), "range"

    if len(numbers) >= 2:
        midpoint = round((numbers[0] + numbers[1]) / 2)
        return company_size_bucket(midpoint), "range"

    bucket = company_size_bucket(numbers[0])
    return (bucket, "count") if bucket else (original, "")


def fuzzy_job_title(value, maps):
    original = str(value or "").strip()
    if not original:
        return "", ""

    exact = maps["jobTitle"].get(normalize_lookup_key(original))
    if exact:
        return exact, "exact"

    lowered = re.sub(r"[^a-z0-9]+", " ", original.lower()).strip()
    for pattern, target in JOB_TITLE_KEYWORD_RULES:
        if re.search(pattern, lowered):
            return target, "keyword"

    candidates = {}
    for item in SCHEMA["jobTitles"]:
        candidates[item["value"]] = item["value"]
        candidates[item["label"]] = item["value"]
    for alias, target in JOB_TITLE_ALIASES.items():
        candidates[alias] = target

    normalized_original = normalize_lookup_key(original)
    best_target = ""
    best_score = 0
    for label, target in candidates.items():
        normalized_label = normalize_lookup_key(label)
        if not normalized_label:
            continue
        score = SequenceMatcher(None, normalized_original, normalized_label).ratio()
        if score > best_score:
            best_score = score
            best_target = target

    if best_score >= 0.88:
        return best_target, "fuzzy"

    if best_score >= 0.78 and best_target in {"Manager", "Director", "Engineering", "Employee", "Student", "Professor"}:
        return best_target, "fuzzy"

    return original, ""


def industry_candidates():
    candidates = {}
    for industry in SCHEMA["industries"]:
        candidates[industry["value"]] = (industry["value"], "")
        candidates[industry["label"]] = (industry["value"], "")
        for sub in industry["subIndustries"]:
            candidates[sub["value"]] = (industry["value"], sub["value"])
            candidates[sub["label"]] = (industry["value"], sub["value"])
    for alias, target in INDUSTRY_ALIASES.items():
        candidates[alias] = target
    return candidates


def match_industry_term(value, maps):
    original = str(value or "").strip()
    if not original:
        return "", "", ""

    normalized = normalize_lookup_key(original)
    industry = maps["industry"].get(normalized)
    if industry:
        return industry, "", "exact"

    sub_industry = maps["subIndustry"].get(normalized)
    if sub_industry:
        return maps["subIndustryParent"].get(sub_industry, ""), sub_industry, "exact"

    alias = INDUSTRY_ALIASES.get(normalize_lookup_key(original)) or INDUSTRY_ALIASES.get(original.lower().strip())
    if alias:
        return alias[0], alias[1], "alias"

    lowered = re.sub(r"[^a-z0-9&+]+", " ", original.lower()).strip()
    for pattern, target in INDUSTRY_KEYWORD_RULES:
        if re.search(pattern, lowered):
            return target[0], target[1], "keyword"

    best_industry = ""
    best_sub = ""
    best_score = 0
    for label, target in industry_candidates().items():
        normalized_label = normalize_lookup_key(label)
        if not normalized_label:
            continue
        score = SequenceMatcher(None, normalized, normalized_label).ratio()
        if score > best_score:
            best_score = score
            best_industry, best_sub = target

    if best_score >= 0.88:
        return best_industry, best_sub, "fuzzy"

    return original, "", ""


def clean_industry_fields(cleaned, maps, warnings):
    original_industry = cleaned["industry"]
    original_sub = cleaned["subIndustry"]

    industry, sub_industry, method = match_industry_term(original_industry, maps)
    if industry and industry != original_industry:
        warnings["industry"] = f"Cleaned industry '{original_industry}' to '{industry}' by {method} match"
    cleaned["industry"] = industry

    sub_from_sub_field = ""
    sub_method = ""
    if original_sub:
        matched_industry, matched_sub, sub_method = match_industry_term(original_sub, maps)
        if matched_sub:
            sub_from_sub_field = matched_sub
            if not cleaned["industry"] or cleaned["industry"] != matched_industry:
                if cleaned["industry"] and cleaned["industry"] != matched_industry:
                    warnings["industry"] = f"Adjusted industry from '{cleaned['industry']}' to '{matched_industry}' based on sub industry '{original_sub}'"
                cleaned["industry"] = matched_industry
        elif cleaned["industry"]:
            scoped = maps["subIndustryByIndustry"].get(cleaned["industry"], {}).get(normalize_lookup_key(original_sub))
            if scoped:
                sub_from_sub_field = scoped
                sub_method = "exact"

    final_sub = sub_from_sub_field or sub_industry
    if final_sub:
        if final_sub != original_sub:
            source = original_sub or original_industry
            warnings["subIndustry"] = f"Cleaned sub industry '{source}' to '{final_sub}' by {(sub_method or method)} match"
        cleaned["subIndustry"] = final_sub

    return cleaned


def clean_upload_fields(fields, missing_columns):
    maps = enum_lookup_maps()
    cleaned = {key: str(fields.get(key, "")).strip() for key in FIELD_TO_HEADER}
    warnings = {}

    if cleaned["workEmail"]:
        lowered = cleaned["workEmail"].replace("mailto:", "").strip().lower()
        if lowered != cleaned["workEmail"]:
            warnings["workEmail"] = f"Cleaned email '{cleaned['workEmail']}' to '{lowered}'"
        cleaned["workEmail"] = lowered

    for field in ["country"]:
        cleaned[field] = map_enum_value(field, cleaned[field], maps, warnings)

    company_size, company_size_method = clean_company_size(cleaned["companySize"], maps)
    if company_size != cleaned["companySize"]:
        warnings["companySize"] = f"Cleaned company size '{cleaned['companySize']}' to '{company_size}' by {company_size_method} match"
    cleaned["companySize"] = company_size

    job_title, job_title_method = fuzzy_job_title(cleaned["jobTitle"], maps)
    if job_title != cleaned["jobTitle"]:
        warnings["jobTitle"] = f"Cleaned job title '{cleaned['jobTitle']}' to '{job_title}' by {job_title_method} match"
    cleaned["jobTitle"] = job_title

    cleaned = clean_industry_fields(cleaned, maps, warnings)

    for header in missing_columns:
        field = next((key for key, value in FIELD_TO_HEADER.items() if value == header), None)
        if field and field not in REQUIRED_FIELDS:
            warnings[field] = f"Missing column '{header}'; treated as blank"

    return cleaned, warnings


def warnings_from_errors(errors):
    return {
        field: f"{FIELD_LABELS.get(field, field)}: {message}" if message == "Required" else message
        for field, message in errors.items()
    }


def duplicate_report(fields, exclude_id=None):
    name_key = normalize_lookup_key(f"{fields.get('firstName', '')} {fields.get('lastName', '')}")
    email = (fields.get("workEmail") or "").strip().lower()
    mobile = re.sub(r"\D+", "", fields.get("mobileNumber") or "")
    leads = load_leads()
    name_matches = []
    email_matches = []
    mobile_matches = []

    for lead in leads:
        if exclude_id and lead["id"] == exclude_id:
            continue
        existing = lead.get("fields", {})
        existing_name_key = normalize_lookup_key(f"{existing.get('firstName', '')} {existing.get('lastName', '')}")
        existing_mobile = re.sub(r"\D+", "", existing.get("mobileNumber") or "")
        if name_key and name_key == existing_name_key:
            name_matches.append(summarize_lead(lead))
        if email and email == (existing.get("workEmail") or "").strip().lower():
            email_matches.append(summarize_lead(lead))
        if mobile and mobile == existing_mobile:
            mobile_matches.append(summarize_lead(lead))

    return {"name": name_matches, "email": email_matches, "mobile": mobile_matches}


def public_duplicate_report(duplicates):
    return {
        "name": bool(duplicates.get("name")),
        "email": bool(duplicates.get("email")),
        "mobile": bool(duplicates.get("mobile")),
    }


def summarize_lead(lead):
    fields = lead.get("fields", {})
    return {
        "id": lead.get("id"),
        "partner": lead.get("partner"),
        "partnerName": lead.get("partnerName"),
        "status": lead.get("status"),
        "firstName": fields.get("firstName"),
        "lastName": fields.get("lastName"),
        "workEmail": fields.get("workEmail"),
        "mobileNumber": fields.get("mobileNumber"),
        "companyName": fields.get("companyName"),
        "createdAt": lead.get("createdAt"),
    }


def validate_lead(fields, block_duplicate_email=True):
    errors = {}
    warnings = {}
    countries, job_titles, company_sizes, industries = schema_sets()

    cleaned = {key: str(fields.get(key, "")).strip() for key in FIELD_TO_HEADER}

    for field in REQUIRED_FIELDS:
        if not cleaned[field]:
            errors[field] = "Required"

    if cleaned["workEmail"] and not EMAIL_RE.match(cleaned["workEmail"]):
        errors["workEmail"] = "Invalid email format"

    if cleaned["country"] and cleaned["country"] not in countries:
        errors["country"] = "Country must match template enum"
    if cleaned["jobTitle"] and cleaned["jobTitle"] not in job_titles:
        errors["jobTitle"] = "Job title must match template enum"
    if cleaned["companySize"] and cleaned["companySize"] not in company_sizes:
        errors["companySize"] = "Company size must match template enum"
    if cleaned["industry"] and cleaned["industry"] not in industries:
        errors["industry"] = "Industry must match template enum"
    if cleaned["industry"] in industries and cleaned["subIndustry"] and cleaned["subIndustry"] not in industries[cleaned["industry"]]:
        errors["subIndustry"] = "Sub industry must belong to selected industry"

    duplicates = duplicate_report(cleaned)
    if duplicates["name"]:
        errors["firstName"] = "This first and last name already exists"
        errors["lastName"] = "This first and last name already exists"
    if duplicates["email"]:
        errors["workEmail"] = "This work email already exists"
    if duplicates["mobile"]:
        errors["mobileNumber"] = "This mobile number already exists"

    return cleaned, errors, warnings, duplicates


def safe_filename_part(value):
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "-", str(value or "").strip())
    return cleaned.strip("-")[:80] or "batch"


def make_export(status, batch_id="", export_tracking_code=""):
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    leads = load_leads()
    selected = []
    for lead in leads:
        if status != "all" and lead.get("status") != status:
            continue
        if batch_id:
            lead_batch_id = lead.get("uploadBatchId") or ""
            if batch_id == "__none__":
                if lead_batch_id:
                    continue
            elif lead_batch_id != batch_id:
                continue
        selected.append(lead)

    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    batch_suffix = f"-{safe_filename_part(batch_id)}" if batch_id else ""
    output_path = EXPORT_DIR / f"lead-export-{status}{batch_suffix}-{timestamp}.xlsx"
    shutil.copyfile(TEMPLATE_PATH, output_path)

    wb = load_workbook(output_path)
    ws = wb["Leads"]

    if ws.max_row > 2:
        ws.delete_rows(3, ws.max_row - 2)

    for index, lead in enumerate(selected, start=3):
        fields = lead.get("fields", {})
        for col_index, key in enumerate(FIELD_TO_HEADER, start=1):
            value = export_tracking_code if key == "trackingCode" and export_tracking_code else fields.get(key, "")
            ws.cell(row=index, column=col_index, value=value)

    wb.save(output_path)
    return output_path, len(selected)


def upload_header_candidates(field):
    return [FIELD_TO_HEADER[field], *FIELD_ALIASES.get(field, [])]


def build_upload_header_index(header_cells):
    normalized_headers = {
        normalize_header(value): index
        for index, value in enumerate(header_cells)
        if value not in (None, "")
    }
    header_index = {}
    for field in FIELD_TO_HEADER:
        for candidate in upload_header_candidates(field):
            key = normalize_header(candidate)
            if key in normalized_headers:
                header_index[field] = normalized_headers[key]
                break
    return header_index


def looks_like_instruction_row(row):
    values = [str(value or "").strip().lower() for value in row if value not in (None, "")]
    if not values:
        return False
    joined = " ".join(values)
    has_email = any(EMAIL_RE.match(value) for value in values)
    instruction_tokens = ["notice", "required", "sample", "example", "do not delete", "note"]
    repeated_placeholder = len(set(values)) == 1 and values[0] in {"note", "sample", "example", "required"}
    return not has_email and (repeated_placeholder or any(token in joined for token in instruction_tokens))


def parse_uploaded_workbook(file_bytes):
    workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws = workbook["Leads"] if "Leads" in workbook.sheetnames else workbook.active
    header_cells = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])
    headers = build_upload_header_index(header_cells)
    recognized = [FIELD_TO_HEADER[field] for field in FIELD_TO_HEADER if field in headers]
    missing = [FIELD_TO_HEADER[field] for field in FIELD_TO_HEADER if field not in headers]
    if not recognized:
        raise ValueError("No matching lead columns found. Keep at least one template header in row 1.")

    row2 = [value for value in next(ws.iter_rows(min_row=2, max_row=2, values_only=True), [])]
    start_row = 3 if looks_like_instruction_row(row2) else 2

    parsed = []
    for row_number, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=start_row):
        fields = {}
        has_value = False
        for field, header in FIELD_TO_HEADER.items():
            if field not in headers:
                fields[field] = ""
                continue
            value = row[headers[field]] if headers[field] < len(row) else None
            if value not in (None, ""):
                has_value = True
            fields[field] = "" if value is None else str(value).strip()
        if has_value:
            parsed.append({"row": row_number, "fields": fields})
    return parsed, missing


def import_uploaded_leads(file_bytes, partner, partner_name=""):
    parsed_rows, missing_columns = parse_uploaded_workbook(file_bytes)
    leads = load_leads()
    imported = []
    failed = []
    row_warnings = []
    seen_names = set()
    seen_emails = set()
    seen_mobiles = set()
    batch_created_at = utc_now()
    batch_id = f"batch-{datetime.utcnow().strftime('%Y%m%d%H%M%S')}-{uuid.uuid4().hex[:6]}"
    batch_label = f"{partner or 'Unknown activity'} · {batch_created_at}"

    for item in parsed_rows:
        cleaned, cleanup_warnings = clean_upload_fields(item["fields"], missing_columns)
        cleaned, errors, validation_warnings, duplicates = validate_lead(cleaned, block_duplicate_email=False)
        warnings = {**cleanup_warnings, **validation_warnings}
        name_key = normalize_lookup_key(f"{cleaned.get('firstName', '')} {cleaned.get('lastName', '')}")
        email_key = cleaned.get("workEmail", "").lower()
        mobile_key = re.sub(r"\D+", "", cleaned.get("mobileNumber", ""))
        if name_key and name_key in seen_names:
            errors["firstName"] = "Duplicate first and last name inside uploaded file"
            errors["lastName"] = "Duplicate first and last name inside uploaded file"
        if email_key and email_key in seen_emails:
            errors["workEmail"] = "Duplicate email inside uploaded file"
        if mobile_key and mobile_key in seen_mobiles:
            errors["mobileNumber"] = "Duplicate mobile number inside uploaded file"

        if name_key:
            seen_names.add(name_key)
        if email_key:
            seen_emails.add(email_key)
        if mobile_key:
            seen_mobiles.add(mobile_key)
        if warnings:
            row_warnings.append(
                {
                    "row": item["row"],
                    "email": cleaned.get("workEmail", ""),
                    "companyName": cleaned.get("companyName", ""),
                    "warnings": warnings,
                }
            )
        if errors:
            failed.append(
                {
                    "row": item["row"],
                    "email": cleaned.get("workEmail", ""),
                    "companyName": cleaned.get("companyName", ""),
                    "errors": errors,
                }
            )
            continue
        lead = {
            "id": uuid.uuid4().hex,
            "partner": partner,
            "partnerName": partner_name,
            "uploadBatchId": batch_id,
            "uploadBatchLabel": batch_label,
            "uploadBatchCreatedAt": batch_created_at,
            "status": "pending",
            "createdAt": utc_now(),
            "updatedAt": utc_now(),
            "reviewNote": "",
            "warnings": warnings,
            "fields": cleaned,
        }
        imported.append(lead)

    if imported:
        save_leads(imported + leads)

    return {
        "parsed": len(parsed_rows),
        "imported": len(imported),
        "failed": len(failed),
        "batchId": batch_id if imported else "",
        "batchLabel": batch_label if imported else "",
        "missingColumns": missing_columns,
        "warnings": row_warnings[:50],
        "errors": failed[:50],
        "partial": bool(imported and failed),
        "rejected": bool(failed and not imported),
    }


class LeadPortalHandler(BaseHTTPRequestHandler):
    server_version = "PartnerLeadPortal/0.1"

    def log_message(self, fmt, *args):
        print("[%s] %s" % (self.log_date_time_string(), fmt % args))

    def send_json(self, payload, status=HTTPStatus.OK):
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_cors_headers()
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def send_cors_headers(self):
        if not CORS_ALLOW_ORIGIN:
            return
        self.send_header("Access-Control-Allow-Origin", CORS_ALLOW_ORIGIN)
        self.send_header("Access-Control-Allow-Methods", "GET, POST, PATCH, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type, X-Admin-Key")
        self.send_header("Access-Control-Max-Age", "86400")

    def read_json(self):
        length = int(self.headers.get("Content-Length", "0"))
        if length == 0:
            return {}
        raw = self.rfile.read(length)
        return json.loads(raw.decode("utf-8"))

    def read_multipart_upload(self):
        length = int(self.headers.get("Content-Length", "0"))
        if length > MAX_UPLOAD_BYTES:
            raise ValueError("File is too large. Maximum upload size is 8 MB.")
        environ = {
            "REQUEST_METHOD": "POST",
            "CONTENT_TYPE": self.headers.get("Content-Type", ""),
            "CONTENT_LENGTH": str(length),
        }
        form = cgi.FieldStorage(fp=self.rfile, headers=self.headers, environ=environ)
        upload = form["file"] if "file" in form else None
        if upload is None or not getattr(upload, "filename", ""):
            raise ValueError("No Excel file uploaded.")
        if not upload.filename.lower().endswith(".xlsx"):
            raise ValueError("Only .xlsx files are supported.")
        partner = form.getfirst("partner", "").strip()
        if not partner:
            raise ValueError("Activity name is required.")
        partner_name = form.getfirst("partnerName", "").strip()
        if not partner_name:
            raise ValueError("Partner name is required.")
        return upload.file.read(), partner, partner_name

    def is_admin_request(self):
        parsed = urlparse(self.path)
        query = parse_qs(parsed.query)
        provided = self.headers.get("X-Admin-Key") or query.get("admin_key", [""])[0]
        return provided == ADMIN_KEY

    def require_admin(self):
        if ADMIN_AUTH_DISABLED:
            return True
        if self.is_admin_request():
            return True
        self.send_json({"error": "Admin access required"}, HTTPStatus.UNAUTHORIZED)
        return False

    def serve_static(self, parsed):
        path = parsed.path
        if PUBLIC_ONLY and path == "/admin":
            self.send_error(HTTPStatus.NOT_FOUND)
            return
        if path in {"/", "/partner", "/admin"}:
            path = "/index.html"
        file_path = (STATIC_DIR / path.lstrip("/")).resolve()
        if not str(file_path).startswith(str(STATIC_DIR.resolve())) or not file_path.exists():
            self.send_error(HTTPStatus.NOT_FOUND)
            return

        content_type = "text/plain; charset=utf-8"
        if file_path.suffix == ".html":
            content_type = "text/html; charset=utf-8"
        elif file_path.suffix == ".css":
            content_type = "text/css; charset=utf-8"
        elif file_path.suffix == ".js":
            content_type = "application/javascript; charset=utf-8"

        body = file_path.read_bytes()
        self.send_response(HTTPStatus.OK)
        self.send_cors_headers()
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def do_OPTIONS(self):
        self.send_response(HTTPStatus.NO_CONTENT)
        self.send_cors_headers()
        self.end_headers()

    def do_GET(self):
        parsed = urlparse(self.path)
        query = parse_qs(parsed.query)

        if parsed.path == "/api/schema":
            self.send_json(SCHEMA)
            return

        if parsed.path == "/healthz":
            self.send_json({"status": "ok"})
            return

        if parsed.path == "/template":
            if not TEMPLATE_PATH.exists():
                self.send_error(HTTPStatus.NOT_FOUND)
                return
            body = TEMPLATE_PATH.read_bytes()
            self.send_response(HTTPStatus.OK)
            self.send_cors_headers()
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", f'attachment; filename="{TEMPLATE_PATH.name}"')
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
            return

        if parsed.path == "/api/leads":
            if PUBLIC_ONLY:
                self.send_error(HTTPStatus.NOT_FOUND)
                return
            if not self.require_admin():
                return
            leads = load_leads()
            self.send_json({"leads": leads})
            return

        if parsed.path == "/api/check":
            fields = {
                "firstName": query.get("firstName", [""])[0],
                "lastName": query.get("lastName", [""])[0],
                "workEmail": query.get("workEmail", [""])[0],
                "mobileNumber": query.get("mobileNumber", [""])[0],
            }
            self.send_json({"duplicates": public_duplicate_report(duplicate_report(fields))})
            return

        if parsed.path == "/api/export":
            if PUBLIC_ONLY:
                self.send_error(HTTPStatus.NOT_FOUND)
                return
            if not self.require_admin():
                return
            status = query.get("status", ["approved"])[0]
            if status not in {"approved", "pending", "rejected", "all"}:
                self.send_json({"error": "Invalid export status"}, HTTPStatus.BAD_REQUEST)
                return
            batch_id = query.get("batchId", [""])[0]
            export_tracking_code = query.get("trackingCode", [""])[0].strip()
            output_path, count = make_export(status, batch_id, export_tracking_code)
            self.send_json({"path": str(output_path), "count": count})
            return

        if parsed.path.startswith("/exports/"):
            if PUBLIC_ONLY:
                self.send_error(HTTPStatus.NOT_FOUND)
                return
            if not self.require_admin():
                return
            file_path = (EXPORT_DIR / parsed.path.removeprefix("/exports/")).resolve()
            if not str(file_path).startswith(str(EXPORT_DIR.resolve())) or not file_path.exists():
                self.send_error(HTTPStatus.NOT_FOUND)
                return
            body = file_path.read_bytes()
            self.send_response(HTTPStatus.OK)
            self.send_cors_headers()
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", f'attachment; filename="{file_path.name}"')
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
            return

        self.serve_static(parsed)

    def do_POST(self):
        parsed = urlparse(self.path)

        if parsed.path == "/api/leads":
            payload = self.read_json()
            partner = str(payload.get("partner", "")).strip()
            partner_name = str(payload.get("partnerName", "")).strip()
            if not partner_name:
                self.send_json({"errors": {"partnerName": "Partner name is required"}}, HTTPStatus.BAD_REQUEST)
                return
            if not partner:
                self.send_json({"errors": {"partner": "Activity name is required"}}, HTTPStatus.BAD_REQUEST)
                return
            cleaned, errors, warnings, duplicates = validate_lead(payload.get("fields", {}))
            if errors:
                self.send_json(
                    {
                        "errors": errors,
                        "warnings": warnings,
                        "duplicates": public_duplicate_report(duplicates),
                    },
                    HTTPStatus.BAD_REQUEST,
                )
                return

            leads = load_leads()
            lead = {
                "id": uuid.uuid4().hex,
                "partner": partner,
                "partnerName": partner_name,
                "status": "pending",
                "createdAt": utc_now(),
                "updatedAt": utc_now(),
                "reviewNote": "",
                "warnings": warnings,
                "fields": cleaned,
            }
            leads.insert(0, lead)
            save_leads(leads)
            self.send_json({"lead": lead, "warnings": warnings}, HTTPStatus.CREATED)
            return

        if parsed.path == "/api/upload":
            try:
                file_bytes, partner, partner_name = self.read_multipart_upload()
                result = import_uploaded_leads(file_bytes, partner, partner_name)
            except Exception as exc:
                self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)
                return
            self.send_json(result, HTTPStatus.BAD_REQUEST if result.get("rejected") else HTTPStatus.CREATED)
            return

        self.send_error(HTTPStatus.NOT_FOUND)

    def do_PATCH(self):
        parsed = urlparse(self.path)
        if not parsed.path.startswith("/api/leads/"):
            self.send_error(HTTPStatus.NOT_FOUND)
            return
        if PUBLIC_ONLY:
            self.send_error(HTTPStatus.NOT_FOUND)
            return
        if not self.require_admin():
            return

        lead_id = parsed.path.split("/")[-1]
        payload = self.read_json()
        leads = load_leads()
        for lead in leads:
            if lead["id"] == lead_id:
                status = payload.get("status")
                if status:
                    if status not in {"pending", "approved", "rejected"}:
                        self.send_json({"error": "Invalid status"}, HTTPStatus.BAD_REQUEST)
                        return
                    lead["status"] = status
                if "reviewNote" in payload:
                    lead["reviewNote"] = str(payload.get("reviewNote") or "").strip()
                lead["updatedAt"] = utc_now()
                save_leads(leads)
                self.send_json({"lead": lead})
                return

        self.send_json({"error": "Lead not found"}, HTTPStatus.NOT_FOUND)


def main():
    host = os.environ.get("HOST", "127.0.0.1")
    port = int(os.environ.get("PORT", "8787"))
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    server = ThreadingHTTPServer((host, port), LeadPortalHandler)
    print(f"Partner Lead Portal running at http://{host}:{port}")
    print(f"Template: {TEMPLATE_PATH}")
    if PUBLIC_ONLY:
        print("Public-only mode: admin routes and export APIs are disabled")
    server.serve_forever()


if __name__ == "__main__":
    main()
